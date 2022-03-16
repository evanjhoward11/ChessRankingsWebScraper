import requests, openpyxl, pandas
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter

# Add a user agent so I don't get blocked from websites
# code is generic, copy/paste it in all web scraping programs
headers = requests.utils.default_headers()
headers.update({
    'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:52.0) Gecko/20100101 Firefox/52.0',
})

url = "https://2700chess.com"
data_cols = ["ranks", "names", "countries", "ratings", "ages"]
ranks = ["#"]
names = ["Name"]
ratings = ["Rating"]
countries = ["Country"]
ages = ["Age"]
col_widths = {"#": 2.29, "Name": 15.86, "Rating": 6.29, "Country": 21.71, "Age": 3.71}

# Initiate typical web scraping and bs4 stuff
res = requests.get(url, headers=headers)
res_text = res.text
soup = BeautifulSoup(res_text, 'lxml')

# Get date text, strip out extra info, and reformat order
date_html = soup.find("strong")
date_full_text = date_html.text
date_split1 = date_full_text.split(",")
date_split2 = date_split1[0].split()
date = f"{date_split2[-2]} {date_split2[-3]}, {date_split2[-1]}"
#print(date)

# Get data from rankings table (rank, name, country, rating, age)
# look at whole table first, then drill down to single player (rows), then collect individual pieces of info (columns) of each player
table = soup.find("tbody", class_="list")
players = table.find_all("tr")
for attribute in players:
    rank_text = attribute.find("td", class_="live_pos live_standard_pos text-standard").text
    rank = rank_text.strip()
    ranks.append(int(rank))
    
    name_class = attribute.find("td", class_="name")
    name = name_class.find("span", class_="hidden searched")
    names.append(name.text)
    
    rating = attribute.find("strong")
    ratings.append(float(rating.text))
    
    country_class = attribute.find("td", class_="country")
    country = country_class.find("span", class_="hidden searched")
    countries.append(country.text)
    
    age_class = attribute.find("td", class_="age")
    age = age_class.find("span")
    ages.append(int(age.text))
    

data = [ranks, names, ratings, countries, ages]


wb = openpyxl.load_workbook("Chess Rankings Scraper.xlsx")
sheet1 = wb.worksheets[0]
sheet1.title = date


for rows in range(len(data[0])):
    for cols in range(len(data)):
        sheet1.cell(row=rows+1, column=cols+1).value = data[cols][rows]

for columns in range(len(data)):
    x = sheet1.cell(row=1, column=columns+1)
    sheet1.column_dimensions[get_column_letter(x.column)].width = col_widths[x.value]








wb.save("Chess Rankings Scraper.xlsx")


# export data to pandas
# highlight cell green or red based on rating change